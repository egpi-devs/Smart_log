import json
import bcrypt
import pyodbc
from datetime import datetime, date

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from django.db import connection
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
import io

<<<<<<< HEAD
<<<<<<< HEAD
# adding sql queries here
=======

>>>>>>> a9a5fba (Smart Log inti)
=======

>>>>>>> a9a5fba (Smart Log inti)
def clean_date(val):
    """Sanitize a date value for SQL Server: empty/falsy → None, strip 'T'."""
    if not val or not str(val).strip():
        return None
    return str(val).strip().replace('T', ' ')


def clean_datetime(val):
    """Alias for clean_date — works for both date and datetime columns."""
    return clean_date(val)


def normalize_datetime_str(val):
    """Normalize a datetime string to 'YYYY-MM-DD HH:MM' for comparison.
    Strips seconds and handles both 'T' and ' ' separators."""
    if not val:
        return ''
    s = str(val).strip().replace('T', ' ')
    # Keep only up to minutes (first 16 chars covers 'YYYY-MM-DD HH:MM')
    return s[:16]


# ──────────────────────────────────────────────
#  HELPERS
# ──────────────────────────────────────────────

def dictfetchall(cursor):
    """Return all rows from a cursor as a list of dicts."""
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def dictfetchone(cursor):
    columns = [col[0] for col in cursor.description]
    row = cursor.fetchone()
    return dict(zip(columns, row)) if row else None


def get_user_roles(cursor, username):
    """Get all roles for a user (primary + extra)."""
    cursor.execute(
        "SELECT Role FROM Users WHERE Username = %s",
        [username]
    )
    row = cursor.fetchone()
    roles = [row[0]] if row else []

    try:
        cursor.execute(
            "SELECT extra_role FROM user_extra_roles WHERE username = %s",
            [username]
        )
        for r in cursor.fetchall():
            if r[0] not in roles:
                roles.append(r[0])
    except Exception:
        pass  # Table might not exist yet

    return roles


def has_any_role(user_roles, *targets):
    return any(r in targets for r in user_roles)


def role_starts_with(user_roles, *prefixes):
    return any(r.startswith(p) for r in user_roles for p in prefixes)


def log_audit(cursor, username, action_type, entity_type, entity_id, old_values=None, new_values=None, details=None):
    """
    Unified logging helper for the new AuditLogs table.
    Old/New values should be JSON strings or None.
    Superusers are stealthy: their actions are not logged here.
    """
    if str(username).lower() == 'superuser':
        return

    cursor.execute("""
        INSERT INTO AuditLogs (
            Timestamp, Username, ActionType, EntityType, EntityID, OldValues, NewValues, Details
        ) VALUES (GETDATE(), %s, %s, %s, %s, %s, %s, %s)
    """, [username, action_type, entity_type, entity_id, old_values, new_values, details])


# ──────────────────────────────────────────────
#  AUTH VIEWS
# ──────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    username = request.data.get('username', '').strip()
    password = request.data.get('password', '').strip()

    if not username or not password:
        return Response({'error': 'Username and password required'},
                        status=status.HTTP_400_BAD_REQUEST)

    with connection.cursor() as cursor:
        cursor.execute("SELECT ID, Password FROM Users WHERE Username = %s", [username])
        row = cursor.fetchone()
        if not row:
            return Response({'error': 'Invalid credentials'},
                            status=status.HTTP_401_UNAUTHORIZED)

        user_id = row[0]
        stored_hash = row[1]
        if not bcrypt.checkpw(password.encode('utf-8'), stored_hash.encode('utf-8')):
            return Response({'error': 'Invalid credentials'},
                            status=status.HTTP_401_UNAUTHORIZED)

        roles = get_user_roles(cursor, username)
        log_audit(
            cursor=cursor,
            username=username,
            action_type='Login',
            entity_type='User',
            entity_id=username,
            details='User logged in'
        )

    # Generate JWT tokens
    from rest_framework_simplejwt.tokens import RefreshToken

    class FakeUser:
        pk = user_id
        id = user_id

    fake = FakeUser()
    refresh = RefreshToken.for_user(fake)
    refresh['username'] = username
    refresh['roles'] = roles

    return Response({
        'access': str(refresh.access_token),
        'refresh': str(refresh),
        'username': username,
        'roles': roles,
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def forgot_password(request):
    username = request.data.get('username', '').strip()
    new_password = request.data.get('new_password', '').strip()
    confirm_password = request.data.get('confirm_password', '').strip()
    admin_username = request.data.get('admin_username', '').strip()
    admin_password = request.data.get('admin_password', '').strip()

    if not all([username, new_password, confirm_password, admin_username, admin_password]):
        return Response({'error': 'All fields are required'}, status=400)

    if new_password != confirm_password:
        return Response({'error': 'Passwords do not match'}, status=400)

    with connection.cursor() as cursor:
        # Validate admin
        cursor.execute(
            "SELECT Password FROM Users WHERE Username = %s AND Role = 'Manager'",
            [admin_username]
        )
        admin_row = cursor.fetchone()
        if not admin_row:
            return Response({'error': 'Invalid admin credentials'}, status=403)

        if not bcrypt.checkpw(admin_password.encode('utf-8'),
                              admin_row[0].encode('utf-8')):
            return Response({'error': 'Invalid admin credentials'}, status=403)

        # Check user exists
        cursor.execute("SELECT COUNT(*) FROM Users WHERE Username = %s", [username])
        if cursor.fetchone()[0] == 0:
            return Response({'error': 'Username does not exist'}, status=404)

        # Update password
        hashed = bcrypt.hashpw(new_password.encode('utf-8'),
                               bcrypt.gensalt()).decode('utf-8')
        cursor.execute("UPDATE Users SET Password = %s WHERE Username = %s",
                       [hashed, username])

    return Response({'message': 'Password reset successfully'})


# ──────────────────────────────────────────────
#  RAW MATERIALS
# ──────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def raw_materials(request):
<<<<<<< HEAD
<<<<<<< HEAD
    with connection.cursor() as cursor:
        roles = get_user_roles(cursor, request.user.username)
        if has_any_role(roles, 'QA Manager') and not has_any_role(roles, 'Manager', 'Superuser'):
             return Response({'error': 'Unauthorized'}, status=403)
=======
>>>>>>> a9a5fba (Smart Log inti)
=======
>>>>>>> a9a5fba (Smart Log inti)
    current_year = datetime.now().strftime('%y')
    qc_pattern = f'%{current_year}RM'

    if request.method == 'GET':
        offset = int(request.query_params.get('offset', 0))
        limit = int(request.query_params.get('limit', 30))

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) FROM RawMaterialData WHERE [QC No.] LIKE %s",
                [qc_pattern]
            )
            total = cursor.fetchone()[0]

            cursor.execute("""
                SELECT r.Date, r.[Material Name], r.[Material Code],
                       p.[ProductCategory], r.[Batch No.], r.[QC No.],
                       r.Micro, r.MicroStatus, r.Chemical, r.ChemicalStatus,
                       r.Manufacturer, r.Supplier,
                       r.[Manufacture Date], r.[Expiry Date], r.Status,
                       r.Notes, r.[Created By], r.[Edited By], r.[Edited Date]
                FROM RawMaterialData r
                JOIN ProductNames p ON r.[Material Name] = p.[ProductName]
                WHERE r.[QC No.] LIKE %s
                ORDER BY r.[QC No.] DESC
                OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
            """, [qc_pattern, offset, limit])
            rows = dictfetchall(cursor)

        return Response({'total': total, 'results': rows})

    elif request.method == 'POST':
        data = request.data
        username = data.get('username', 'Unknown')
        user_roles = data.get('user_roles', [])

        micro_s = data.get('micro_status', 'Pending')
        chem_s = data.get('chemical_status', 'Pending')
        if micro_s.lower() == 'accepted' and chem_s.lower() == 'accepted':
            final_status = 'Accepted'
        elif 'rejected' in (micro_s.lower(), chem_s.lower()):
            final_status = 'Rejected'
        else:
            final_status = 'Pending'

        with connection.cursor() as cursor:
            # Superuser stealth: masquerade as the creator of the previous row
            actual_creator = username
            if str(username).lower() == 'superuser':
                cursor.execute("SELECT TOP 1 [Created By] FROM RawMaterialData ORDER BY Date DESC, [QC No.] DESC")
                prev_creator_row = cursor.fetchone()
                if prev_creator_row and prev_creator_row[0]:
                    actual_creator = prev_creator_row[0]
                else:
                    actual_creator = 'System'

            cursor.execute("""
                INSERT INTO RawMaterialData (
                    Date, [Material Name], [Material Code], [Material Category],
                    [Batch No.], [QC No.], Micro, MicroStatus, Chemical, ChemicalStatus,
                    Manufacturer, Supplier, [Manufacture Date], [Expiry Date], Status,
                    Notes, [Created By]
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, [
                clean_datetime(data['date']), data['material'], data['material_code'],
                data.get('category', ''), data['batch'], data['qc'],
                data.get('micro', False), micro_s,
                data.get('chemical', False), chem_s,
                data.get('manufacturer', ''), data.get('supplier', ''),
                clean_date(data.get('manufacture_date')), clean_date(data.get('expiry_date')),
                final_status, data.get('notes', ''), actual_creator
            ])

            # Also insert into unified audit log
            log_audit(
                cursor=cursor,
                username=username,
                action_type='Creation',
                entity_type='Raw Material',
                entity_id=data['qc'],
                details=f"Status: {data.get('status', 'Pending')}"
            )

        return Response({'message': 'Entry added successfully'}, status=201)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def raw_material_detail(request, qc_no):
    data = request.data
    username = data.get('username', 'Unknown')
    user_roles = data.get('user_roles', [])
    today = date.today().isoformat()

    micro_s = data.get('micro_status', 'Pending')
    chem_s = data.get('chemical_status', 'Pending')
    if micro_s.lower() == 'accepted' and chem_s.lower() == 'accepted':
        final_status = 'Accepted'
    elif 'rejected' in (micro_s.lower(), chem_s.lower()):
        final_status = 'Rejected'
    else:
        final_status = 'Pending'

    with connection.cursor() as cursor:
        # Get old data for audit
        cursor.execute("""
            SELECT Date, [Material Name], [Material Code], [Material Category],
                   [Batch No.], [QC No.], Micro, MicroStatus, Chemical, ChemicalStatus,
                   Manufacturer, Supplier, [Manufacture Date], [Expiry Date], Status, Notes
            FROM RawMaterialData WHERE [QC No.] = %s
        """, [qc_no])
        old_row = cursor.fetchone()
        if not old_row:
            return Response({'error': 'Record not found'}, status=404)

        old_data = {
            'date': str(old_row[0]), 'material': old_row[1],
            'material_code': old_row[2], 'category': old_row[3],
            'batch': old_row[4], 'qc': old_row[5],
            'micro': old_row[6], 'micro_status': old_row[7] or 'Pending',
            'chemical': old_row[8], 'chemical_status': old_row[9] or 'Pending',
            'manufacturer': old_row[10] or '', 'supplier': old_row[11] or '',
            'manufacture_date': str(old_row[12] or ''), 'expiry_date': str(old_row[13] or ''),
            'status': old_row[14], 'notes': old_row[15] or ''
        }

        # Use incoming value if supplied; fall back to existing DB value
        new_mfg  = clean_date(data.get('manufacture_date')) if data.get('manufacture_date') else old_row[12]
        new_exp  = clean_date(data.get('expiry_date'))      if data.get('expiry_date')      else old_row[13]

        # Check if creation date is being edited
        is_superuser = (str(username).lower() == 'superuser')
        old_date_str = str(clean_datetime(old_data['date']))
        new_date_str = str(clean_datetime(data['date']))
        date_changed = normalize_datetime_str(old_date_str) != normalize_datetime_str(new_date_str)

        # If superuser changes date: clear audit history and remove edited traces
        if is_superuser and date_changed:
            cursor.execute("DELETE FROM AuditLogs WHERE EntityID = %s", [qc_no])
            
            cursor.execute("""
                UPDATE RawMaterialData
                SET Date=%s, [Material Name]=%s, [Material Code]=%s,
                    [Material Category]=%s, [Batch No.]=%s, [QC No.]=%s,
                    Micro=%s, MicroStatus=%s, Chemical=%s, ChemicalStatus=%s,
                    Manufacturer=%s, Supplier=%s,
                    [Manufacture Date]=%s, [Expiry Date]=%s, Status=%s,
                    Notes=%s, [Edited By]=NULL, [Edited Date]=NULL
                WHERE [QC No.] = %s
            """, [
                new_date_str, data['material'], data['material_code'],
                data.get('category', ''), data['batch'], data['qc'],
                data.get('micro', False), micro_s,
                data.get('chemical', False), chem_s,
                data.get('manufacturer', ''), data.get('supplier', ''),
                new_mfg, new_exp,
                final_status, data.get('notes', ''), qc_no
            ])
            return Response({'message': 'Entry date updated and audit cleared successfully'})

        # Find other changes
        changes = {}
        compare_data = {
            **data,
            'manufacture_date': new_mfg or '',
            'expiry_date': new_exp or '',
            'micro': data.get('micro', False),
            'micro_status': micro_s,
            'chemical': data.get('chemical', False),
            'chemical_status': chem_s,
            'manufacturer': data.get('manufacturer', ''),
            'supplier': data.get('supplier', ''),
            'status': final_status,
            'notes': data.get('notes', '')
        }
        for key in old_data:
            if key == 'date':
                continue  # date changes are handled separately above
            if key in compare_data and str(compare_data[key]) != str(old_data[key]):
                changes[key] = {'old': str(old_data[key]), 'new': str(compare_data[key])}

        # Build update query (superusers don't leave Edit traces)
        update_cols = """
            Date=%s, [Material Name]=%s, [Material Code]=%s,
            [Material Category]=%s, [Batch No.]=%s, [QC No.]=%s,
            Micro=%s, MicroStatus=%s, Chemical=%s, ChemicalStatus=%s,
            Manufacturer=%s, Supplier=%s,
            [Manufacture Date]=%s, [Expiry Date]=%s, Status=%s,
            Notes=%s
        """
        update_params = [
            new_date_str, data['material'], data['material_code'],
            data.get('category', ''), data['batch'], data['qc'],
            data.get('micro', False), micro_s,
            data.get('chemical', False), chem_s,
            data.get('manufacturer', ''), data.get('supplier', ''),
            new_mfg, new_exp,
            final_status, data.get('notes', '')
        ]

        if not is_superuser:
            update_cols += ", [Edited By]=%s, [Edited Date]=%s"
            update_params.extend([username, today])

        update_params.append(qc_no)
        
        cursor.execute(f"UPDATE RawMaterialData SET {update_cols} WHERE [QC No.] = %s", update_params)

        if changes:
            if is_superuser:
                # Stealth edit: inject into last audit log if it exists
                cursor.execute("""
                    SELECT TOP 1 LogID, OldValues, NewValues 
                    FROM AuditLogs 
                    WHERE EntityID = %s 
                    ORDER BY Timestamp DESC
                """, [qc_no])
                last_log = dictfetchone(cursor)
                
                if last_log and last_log.get('NewValues'):
                    try:
                        old_vals_dict = json.loads(last_log['OldValues'] or '{}')
                        new_vals_dict = json.loads(last_log['NewValues'] or '{}')
                        for k, v in changes.items():
                            if k not in new_vals_dict:
                                old_vals_dict[k] = v['old']
                            new_vals_dict[k] = v['new']
                            
                        cursor.execute(
                            "UPDATE AuditLogs SET OldValues=%s, NewValues=%s WHERE LogID=%s",
                            [json.dumps(old_vals_dict), json.dumps(new_vals_dict), last_log['LogID']]
                        )
                    except json.JSONDecodeError:
                        pass # Ignore malformed logs
            else:
                # Normal edit: standard log
                log_audit(
                    cursor=cursor,
                    username=username,
                    action_type='Edit',
                    entity_type='Raw Material',
                    entity_id=qc_no,
                    old_values=json.dumps({k: v['old'] for k, v in changes.items()}),
                    new_values=json.dumps({k: v['new'] for k, v in changes.items()})
                )

    return Response({'message': 'Entry updated successfully'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def raw_materials_search(request):
    data = request.data
    current_year = datetime.now().strftime('%y')

    query = """
        SELECT r.Date, r.[Material Name], r.[Material Code],
               p.[ProductCategory], r.[Batch No.], r.[QC No.],
               r.Micro, r.MicroStatus, r.Chemical, r.ChemicalStatus, r.Manufacturer, r.Supplier,
               r.[Manufacture Date], r.[Expiry Date], r.Status,
               r.Notes, r.[Created By], r.[Edited By], r.[Edited Date]
        FROM RawMaterialData r
        JOIN ProductNames p ON r.[Material Name] = p.[ProductName]
        WHERE 1=1
    """
    params = []

    year_filter = data.get('year_filter', current_year)
    query += " AND r.[QC No.] LIKE %s"
    params.append(f'%{year_filter}RM')

    for field, col in [
        ('material_name', "r.[Material Name]"),
        ('batch_number', "r.[Batch No.]"),
        ('qc_number', "r.[QC No.]"),
        ('material_code', "r.[Material Code]"),
        ('manufacturer', "r.Manufacturer"),
        ('supplier', "r.Supplier"),
<<<<<<< HEAD
<<<<<<< HEAD
=======
        ('status', "r.Status"),
>>>>>>> a9a5fba (Smart Log inti)
=======
        ('status', "r.Status"),
>>>>>>> a9a5fba (Smart Log inti)
        ('notes', "r.Notes"),
    ]:
        val = data.get(field, '').strip()
        if val:
            query += f" AND {col} LIKE %s"
            params.append(f'%{val}%')

<<<<<<< HEAD
<<<<<<< HEAD
    # Exact match for status fields
    for field, col in [
        ('status', "r.Status"),
        ('micro_status', "r.MicroStatus"),
        ('chemical_status', "r.ChemicalStatus"),
    ]:
        val = data.get(field, '').strip()
        if val:
            query += f" AND {col} = %s"
            params.append(val)

=======
>>>>>>> a9a5fba (Smart Log inti)
=======
>>>>>>> a9a5fba (Smart Log inti)
    if data.get('product_category'):
        query += " AND p.[ProductCategory] LIKE %s"
        params.append(f'%{data["product_category"]}%')

    if data.get('from_date'):
        query += " AND r.Date >= %s"
        params.append(data['from_date'])
    if data.get('to_date'):
        query += " AND r.Date <= %s"
        params.append(data['to_date'])

    query += " ORDER BY r.[QC No.] DESC"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = dictfetchall(cursor)

    return Response({'results': rows, 'total': len(rows)})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def raw_material_generate_qc(request):
    """Generate next QC number for raw materials."""
    current_year = datetime.now().strftime('%y')
    category_code = 'RM'

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT TOP 1 [QC No.]
            FROM RawMaterialData
            WHERE [Material Category] = 'Raw'
            AND [QC No.] LIKE %s
            ORDER BY [QC No.] DESC
        """, [f'%\\{current_year}{category_code}'])

        latest = cursor.fetchone()

        if latest and latest[0]:
            try:
                qc_str = latest[0]
                seq_part = qc_str.split('\\')[0]
                seq_num = int(seq_part.split('-')[0])
                new_num = f'{seq_num + 1:04d}'
            except (IndexError, ValueError):
                new_num = '0001'
        else:
            new_num = '0001'

        qc_number = f'{new_num}-00\\{current_year}{category_code}'

        # Ensure uniqueness
        cursor.execute(
            "SELECT COUNT(*) FROM RawMaterialData WHERE [QC No.] = %s",
            [qc_number]
        )
        while cursor.fetchone()[0] > 0:
            seq_num = int(new_num)
            new_num = f'{seq_num + 1:04d}'
            qc_number = f'{new_num}-00\\{current_year}{category_code}'
            cursor.execute(
                "SELECT COUNT(*) FROM RawMaterialData WHERE [QC No.] = %s",
                [qc_number]
            )

    return Response({'qc_number': qc_number})


# ──────────────────────────────────────────────
#  PACKAGING MATERIALS
# ──────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def packaging_materials(request):
<<<<<<< HEAD
<<<<<<< HEAD
    with connection.cursor() as cursor:
        roles = get_user_roles(cursor, request.user.username)
        if has_any_role(roles, 'QA Manager') and not has_any_role(roles, 'Manager', 'Superuser'):
             return Response({'error': 'Unauthorized'}, status=403)
=======
>>>>>>> a9a5fba (Smart Log inti)
=======
>>>>>>> a9a5fba (Smart Log inti)
    current_year = datetime.now().strftime('%y')
    qc_pattern = f'%{current_year}PM'

    if request.method == 'GET':
        offset = int(request.query_params.get('offset', 0))
        limit = int(request.query_params.get('limit', 30))

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) FROM PackagingMaterialData WHERE [QC No.] LIKE %s",
                [qc_pattern]
            )
            total = cursor.fetchone()[0]

            cursor.execute("""
                SELECT r.Date, r.[Material Name], r.[Material Code],
                       p.[ProductCategory], r.[QC No.],
                       r.Micro, r.MicroStatus, r.Chemical, r.ChemicalStatus, r.Manufacturer, r.Supplier,
                       r.Status, r.Notes, r.[Created By],
                       r.[Edited By], r.[Edited Date]
                FROM PackagingMaterialData r
                JOIN ProductNames p ON r.[Material Name] = p.[ProductName]
                WHERE r.[QC No.] LIKE %s
                ORDER BY r.[QC No.] DESC
                OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
            """, [qc_pattern, offset, limit])
            rows = dictfetchall(cursor)

        return Response({'total': total, 'results': rows})

    elif request.method == 'POST':
        data = request.data
        username = data.get('username', 'Unknown')
        user_roles = data.get('user_roles', [])

        micro_s = data.get('micro_status', 'Pending')
        chem_s = data.get('chemical_status', 'Pending')
        if micro_s.lower() == 'accepted' and chem_s.lower() == 'accepted':
            final_status = 'Accepted'
        elif 'rejected' in (micro_s.lower(), chem_s.lower()):
            final_status = 'Rejected'
        else:
            final_status = 'Pending'

        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO PackagingMaterialData (
                    Date, [Material Name], [Material Code], [Material Category],
                    [QC No.], Micro, MicroStatus, Chemical, ChemicalStatus, Manufacturer, Supplier,
                    Status, Notes, [Created By]
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, [
                clean_datetime(data['date']), data['material'], data['material_code'],
                data.get('category', ''), data['qc'],
                data.get('micro', False), micro_s,
                data.get('chemical', False), chem_s,
                data.get('manufacturer', ''), data.get('supplier', ''),
                final_status, data.get('notes', ''), username
            ])

            # Also insert into unified audit log
            log_audit(
                cursor=cursor,
                username=username,
                action_type='Creation',
                entity_type='Packaging Material',
                entity_id=data['qc'],
                details=f"Status: {data.get('status', 'Pending')}"
            )

        return Response({'message': 'Entry added successfully'}, status=201)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def packaging_material_detail(request, qc_no):
    data = request.data
    username = data.get('username', 'Unknown')
    user_roles = data.get('user_roles', [])
    today = date.today().isoformat()

    micro_s = data.get('micro_status', 'Pending')
    chem_s = data.get('chemical_status', 'Pending')
    if micro_s.lower() == 'accepted' and chem_s.lower() == 'accepted':
        final_status = 'Accepted'
    elif 'rejected' in (micro_s.lower(), chem_s.lower()):
        final_status = 'Rejected'
    else:
        final_status = 'Pending'

    with connection.cursor() as cursor:
        # Get old data for audit
        cursor.execute("""
            SELECT Date, [Material Name], [Material Code], [Material Category],
                   [QC No.], Micro, MicroStatus, Chemical, ChemicalStatus, Manufacturer, Supplier,
                   Status, Notes
            FROM PackagingMaterialData WHERE [QC No.] = %s
        """, [qc_no])
        old_row = cursor.fetchone()
        if not old_row:
            return Response({'error': 'Record not found'}, status=404)

        old_data = {
            'date': str(old_row[0]), 'material': old_row[1],
            'material_code': old_row[2], 'category': old_row[3],
            'qc': old_row[4], 'micro': old_row[5], 'micro_status': old_row[6] or 'Pending',
            'chemical': old_row[7], 'chemical_status': old_row[8] or 'Pending',
            'manufacturer': old_row[9] or '', 'supplier': old_row[10] or '',
            'status': old_row[11], 'notes': old_row[12] or ''
        }

        # Check if creation date is being edited
        is_superuser = (str(username).lower() == 'superuser')
        old_date_str = str(clean_datetime(old_data['date']))
        new_date_str = str(clean_datetime(data['date']))
        date_changed = normalize_datetime_str(old_date_str) != normalize_datetime_str(new_date_str)

        # If superuser changes date: clear audit history and remove edited traces
        if is_superuser and date_changed:
            cursor.execute("DELETE FROM AuditLogs WHERE EntityID = %s", [qc_no])
            
            cursor.execute("""
                UPDATE PackagingMaterialData
                SET Date=%s, [Material Name]=%s, [Material Code]=%s,
                    [Material Category]=%s, [QC No.]=%s,
                    Micro=%s, MicroStatus=%s, Chemical=%s, ChemicalStatus=%s, Manufacturer=%s, Supplier=%s,
                    Status=%s, Notes=%s, [Edited By]=NULL, [Edited Date]=NULL
                WHERE [QC No.] = %s
            """, [
                new_date_str, data['material'], data['material_code'],
                data.get('category', ''), data['qc'],
                data.get('micro', False), micro_s,
                data.get('chemical', False), chem_s,
                data.get('manufacturer', ''), data.get('supplier', ''),
                final_status, data.get('notes', ''), qc_no
            ])
            return Response({'message': 'Entry date updated and audit cleared successfully'})

        # Find other changes
        changes = {}
        compare_data = {
            **data,
            'micro': data.get('micro', False),
            'micro_status': micro_s,
            'chemical': data.get('chemical', False),
            'chemical_status': chem_s,
            'manufacturer': data.get('manufacturer', ''),
            'supplier': data.get('supplier', ''),
            'status': final_status,
            'notes': data.get('notes', '')
        }
        for key in old_data:
            if key == 'date':
                continue  # date changes are handled separately above
            if key in compare_data and str(compare_data[key]) != str(old_data[key]):
                changes[key] = {'old': str(old_data[key]), 'new': str(compare_data[key])}

        # Build update query (superusers don't leave Edit traces)
        update_cols = """
            Date=%s, [Material Name]=%s, [Material Code]=%s,
            [Material Category]=%s, [QC No.]=%s,
            Micro=%s, MicroStatus=%s, Chemical=%s, ChemicalStatus=%s, Manufacturer=%s, Supplier=%s,
            Status=%s, Notes=%s
        """
        update_params = [
            new_date_str, data['material'], data['material_code'],
            data.get('category', ''), data['qc'],
            data.get('micro', False), micro_s,
            data.get('chemical', False), chem_s,
            data.get('manufacturer', ''), data.get('supplier', ''),
            final_status, data.get('notes', '')
        ]

        if not is_superuser:
            update_cols += ", [Edited By]=%s, [Edited Date]=%s"
            update_params.extend([username, today])

        update_params.append(qc_no)
        
        cursor.execute(f"UPDATE PackagingMaterialData SET {update_cols} WHERE [QC No.] = %s", update_params)

        if changes:
            if is_superuser:
                # Stealth edit: inject into last audit log if it exists
                cursor.execute("""
                    SELECT TOP 1 LogID, OldValues, NewValues 
                    FROM AuditLogs 
                    WHERE EntityID = %s 
                    ORDER BY Timestamp DESC
                """, [qc_no])
                last_log = dictfetchone(cursor)
                
                if last_log and last_log.get('NewValues'):
                    try:
                        old_vals_dict = json.loads(last_log['OldValues'] or '{}')
                        new_vals_dict = json.loads(last_log['NewValues'] or '{}')
                        for k, v in changes.items():
                            if k not in new_vals_dict:
                                old_vals_dict[k] = v['old']
                            new_vals_dict[k] = v['new']
                            
                        cursor.execute(
                            "UPDATE AuditLogs SET OldValues=%s, NewValues=%s WHERE LogID=%s",
                            [json.dumps(old_vals_dict), json.dumps(new_vals_dict), last_log['LogID']]
                        )
                    except json.JSONDecodeError:
                        pass # Ignore malformed logs
            else:
                # Normal edit: standard log
                log_audit(
                    cursor=cursor,
                    username=username,
                    action_type='Edit',
                    entity_type='Packaging Material',
                    entity_id=qc_no,
                    old_values=json.dumps({k: v['old'] for k, v in changes.items()}),
                    new_values=json.dumps({k: v['new'] for k, v in changes.items()})
                )

    return Response({'message': 'Entry updated successfully'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def packaging_materials_search(request):
    data = request.data
    current_year = datetime.now().strftime('%y')

    query = """
        SELECT r.Date, r.[Material Name], r.[Material Code],
               p.[ProductCategory], r.[QC No.],
               r.Micro, r.MicroStatus, r.Chemical, r.ChemicalStatus, r.Manufacturer, r.Supplier,
               r.Status, r.Notes, r.[Created By],
               r.[Edited By], r.[Edited Date]
        FROM PackagingMaterialData r
        JOIN ProductNames p ON r.[Material Name] = p.[ProductName]
        WHERE 1=1
    """
    params = []

    year_filter = data.get('year_filter', current_year)
    query += " AND r.[QC No.] LIKE %s"
    params.append(f'%{year_filter}PM')

    for field, col in [
        ('material_name', "r.[Material Name]"),
        ('qc_number', "r.[QC No.]"),
        ('material_code', "r.[Material Code]"),
        ('manufacturer', "r.Manufacturer"),
        ('supplier', "r.Supplier"),
<<<<<<< HEAD
<<<<<<< HEAD
=======
        ('status', "r.Status"),
>>>>>>> a9a5fba (Smart Log inti)
=======
        ('status', "r.Status"),
>>>>>>> a9a5fba (Smart Log inti)
        ('notes', "r.Notes"),
    ]:
        val = data.get(field, '').strip()
        if val:
            query += f" AND {col} LIKE %s"
            params.append(f'%{val}%')

<<<<<<< HEAD
<<<<<<< HEAD
    # Exact match for status fields
    for field, col in [
        ('status', "r.Status"),
        ('micro_status', "r.MicroStatus"),
        ('chemical_status', "r.ChemicalStatus"),
    ]:
        val = data.get(field, '').strip()
        if val:
            query += f" AND {col} = %s"
            params.append(val)

=======
>>>>>>> a9a5fba (Smart Log inti)
=======
>>>>>>> a9a5fba (Smart Log inti)
    query += " ORDER BY r.[QC No.] DESC"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = dictfetchall(cursor)

    return Response({'results': rows, 'total': len(rows)})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def packaging_material_generate_qc(request):
    current_year = datetime.now().strftime('%y')
    category_code = 'PM'

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT TOP 1 [QC No.]
            FROM PackagingMaterialData
            WHERE [Material Category] = 'Packaged'
            AND [QC No.] LIKE %s
            ORDER BY [QC No.] DESC
        """, [f'%\\{current_year}{category_code}'])

        latest = cursor.fetchone()
        if latest and latest[0]:
            try:
                seq_part = latest[0].split('\\')[0]
                seq_num = int(seq_part.split('-')[0])
                new_num = f'{seq_num + 1:04d}'
            except (IndexError, ValueError):
                new_num = '0001'
        else:
            new_num = '0001'

        qc_number = f'{new_num}-00\\{current_year}{category_code}'

        cursor.execute(
            "SELECT COUNT(*) FROM PackagingMaterialData WHERE [QC No.] = %s",
            [qc_number]
        )
        while cursor.fetchone()[0] > 0:
            seq_num = int(new_num)
            new_num = f'{seq_num + 1:04d}'
            qc_number = f'{new_num}-00\\{current_year}{category_code}'
            cursor.execute(
                "SELECT COUNT(*) FROM PackagingMaterialData WHERE [QC No.] = %s",
                [qc_number]
            )

    return Response({'qc_number': qc_number})


# ──────────────────────────────────────────────
#  FINISHED PRODUCTS
# ──────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def finished_products(request):
<<<<<<< HEAD
<<<<<<< HEAD
    with connection.cursor() as cursor:
        roles = get_user_roles(cursor, request.user.username)
        if has_any_role(roles, 'QA Manager') and not has_any_role(roles, 'Manager', 'Superuser'):
             return Response({'error': 'Unauthorized'}, status=403)
=======
>>>>>>> a9a5fba (Smart Log inti)
=======
>>>>>>> a9a5fba (Smart Log inti)
    current_year = datetime.now().strftime('%y')
    qc_pattern = f'%{current_year}FP'

    if request.method == 'GET':
        offset = int(request.query_params.get('offset', 0))
        limit = int(request.query_params.get('limit', 30))

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) FROM FinishedProductData WHERE [QC No.] LIKE %s",
                [qc_pattern]
            )
            total = cursor.fetchone()[0]

            cursor.execute("""
                SELECT Date, [Material Name], [Material Code],
                       [Material Category], [Batch No.], [QC No.],
                       Micro, MicroStatus, Chemical, ChemicalStatus,
                       [Manufacture Date], [Expiry Date], Status,
                       Notes, [Created By],
                       Reviewed, [Reviewed By], [Reviewed Date]
                FROM FinishedProductData
                WHERE [QC No.] LIKE %s
                ORDER BY [QC No.] DESC
                OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
            """, [qc_pattern, offset, limit])
            rows = dictfetchall(cursor)

        return Response({'total': total, 'results': rows})

    elif request.method == 'POST':
        data = request.data
        username = data.get('username', 'Unknown')

        micro_s = data.get('micro_status', 'Pending')
        chem_s = data.get('chemical_status', 'Pending')
        if micro_s.lower() == 'accepted' and chem_s.lower() == 'accepted':
            final_status = 'Accepted'
        elif 'rejected' in (micro_s.lower(), chem_s.lower()):
            final_status = 'Rejected'
        else:
            final_status = 'Pending'

        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO FinishedProductData (
                    Date, [Material Name], [Material Code], [Material Category],
                    [Batch No.], [QC No.], Micro, MicroStatus,
                    Chemical, ChemicalStatus,
                    [Manufacture Date], [Expiry Date], Status,
                    Notes, [Created By]
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, [
                clean_datetime(data['date']), data['material'], data['material_code'],
                data.get('category', ''), data['batch'], data['qc'],
                data.get('micro', False), micro_s,
                data.get('chemical', False), chem_s,
                clean_date(data.get('manufacture_date')), clean_date(data.get('expiry_date')),
                final_status, data.get('notes', ''), username
            ])

            # Also insert into unified audit log
            log_audit(
                cursor=cursor,
                username=username,
                action_type='Creation',
                entity_type='Finished Product',
                entity_id=data['qc'],
                details=f"Status: {final_status}"
            )

        return Response({'message': 'Entry added successfully'}, status=201)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def finished_product_detail(request, qc_no):
    data = request.data
    username = data.get('username', 'Unknown')

    micro_s = data.get('micro_status', 'Pending')
    chem_s = data.get('chemical_status', 'Pending')
    if micro_s.lower() == 'accepted' and chem_s.lower() == 'accepted':
        final_status = 'Accepted'
    elif 'rejected' in (micro_s.lower(), chem_s.lower()):
        final_status = 'Rejected'
    else:
        final_status = 'Pending'

    with connection.cursor() as cursor:
        # Fetch existing data for audit and fallback dates
        cursor.execute("""
            SELECT Date, [Material Name], [Material Code], [Material Category],
                   [Batch No.], [QC No.], Micro, MicroStatus, Chemical, ChemicalStatus,
                   [Manufacture Date], [Expiry Date], Status, Notes, Reviewed, [Reviewed By], [Reviewed Date]
            FROM FinishedProductData WHERE [QC No.] = %s
        """, [qc_no])
        old_row = cursor.fetchone()
        if not old_row:
            return Response({'error': 'Record not found'}, status=404)

        old_data = {
            'date': str(old_row[0]), 'material': old_row[1],
            'material_code': old_row[2], 'category': old_row[3],
            'batch': old_row[4], 'qc': old_row[5],
            'micro': old_row[6], 'micro_status': old_row[7],
            'chemical': old_row[8], 'chemical_status': old_row[9],
            'manufacture_date': str(old_row[10] or ''), 'expiry_date': str(old_row[11] or ''),
            'status': old_row[12], 'notes': old_row[13]
        }

        # Use incoming value if supplied; fall back to existing DB value
        new_mfg  = clean_date(data.get('manufacture_date')) if data.get('manufacture_date') else old_row[10]
        new_exp  = clean_date(data.get('expiry_date'))      if data.get('expiry_date')      else old_row[11]

        # Check if creation date is being edited
        is_superuser = (str(username).lower() == 'superuser')
        old_date_str = str(clean_datetime(old_data['date']))
        new_date_str = str(clean_datetime(data['date']))
        date_changed = normalize_datetime_str(old_date_str) != normalize_datetime_str(new_date_str)

        # If superuser changes date: clear audit history
        if is_superuser and date_changed:
            cursor.execute("DELETE FROM AuditLogs WHERE EntityID = %s", [qc_no])
            
            cursor.execute("""
                UPDATE FinishedProductData
                SET Date=%s, [Material Name]=%s, [Material Code]=%s,
                    [Material Category]=%s, [Batch No.]=%s, [QC No.]=%s,
                    Micro=%s, MicroStatus=%s, Chemical=%s, ChemicalStatus=%s,
                    [Manufacture Date]=%s, [Expiry Date]=%s, Status=%s, Notes=%s
                WHERE [QC No.] = %s
            """, [
                new_date_str, data['material'], data['material_code'],
                data.get('category', ''), data['batch'], data['qc'],
                data.get('micro', False), micro_s,
                data.get('chemical', False), chem_s,
                new_mfg, new_exp,
                final_status, data.get('notes', ''), qc_no
            ])
            # Note: FinishedProductData doesn't have [Edited By] or [Edited Date] currently on PUT
            return Response({'message': 'Entry date updated and audit cleared successfully'})

        # Find other changes
        changes = {}
        compare_data = {
            **data,
            'micro': data.get('micro', False),
            'micro_status': micro_s,
            'chemical': data.get('chemical', False),
            'chemical_status': chem_s,
            'manufacture_date': new_mfg or '',
            'expiry_date': new_exp or '',
            'status': final_status,
            'notes': data.get('notes', '')
        }
        for key in old_data:
            if key == 'date':
                continue  # date changes are handled separately above
            if key in compare_data and str(compare_data[key]) != str(old_data[key]):
                changes[key] = {'old': str(old_data[key]), 'new': str(compare_data[key])}

        cursor.execute("""
            UPDATE FinishedProductData
            SET Date=%s, [Material Name]=%s, [Material Code]=%s,
                [Material Category]=%s, [Batch No.]=%s, [QC No.]=%s,
                Micro=%s, MicroStatus=%s, Chemical=%s, ChemicalStatus=%s,
                [Manufacture Date]=%s, [Expiry Date]=%s, Status=%s, Notes=%s
            WHERE [QC No.] = %s
        """, [
            new_date_str, data['material'], data['material_code'],
            data.get('category', ''), data['batch'], data['qc'],
            data.get('micro', False), micro_s,
            data.get('chemical', False), chem_s,
            new_mfg, new_exp,
            final_status, data.get('notes', ''), qc_no
        ])

        if changes:
            if is_superuser:
                # Stealth edit: inject into last audit log if it exists
                cursor.execute("""
                    SELECT TOP 1 LogID, OldValues, NewValues 
                    FROM AuditLogs 
                    WHERE EntityID = %s 
                    ORDER BY Timestamp DESC
                """, [qc_no])
                last_log = dictfetchone(cursor)
                
                if last_log and last_log.get('NewValues'):
                    try:
                        old_vals_dict = json.loads(last_log['OldValues'] or '{}')
                        new_vals_dict = json.loads(last_log['NewValues'] or '{}')
                        for k, v in changes.items():
                            if k not in new_vals_dict:
                                old_vals_dict[k] = v['old']
                            new_vals_dict[k] = v['new']
                            
                        cursor.execute(
                            "UPDATE AuditLogs SET OldValues=%s, NewValues=%s WHERE LogID=%s",
                            [json.dumps(old_vals_dict), json.dumps(new_vals_dict), last_log['LogID']]
                        )
                    except json.JSONDecodeError:
                        pass # Ignore malformed logs
            else:
                # Normal edit: standard log
                log_audit(
                    cursor=cursor,
                    username=username,
                    action_type='Edit',
                    entity_type='Finished Product',
                    entity_id=qc_no,
                    old_values=json.dumps({k: v['old'] for k, v in changes.items()}),
                    new_values=json.dumps({k: v['new'] for k, v in changes.items()})
                )

    return Response({'message': 'Entry updated successfully'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def finished_product_mark_reviewed(request, qc_no):
    data = request.data
    username = data.get('username', 'Unknown')
    checked = data.get('chemical', False)
    new_status = data.get('chemical_status', 'Pending')
    today = date.today().isoformat()

    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT Chemical, ChemicalStatus, MicroStatus, Status FROM FinishedProductData WHERE [QC No.] = %s",
            [qc_no]
        )
        old = cursor.fetchone()
        if not old:
            return Response({'error': 'Record not found'}, status=404)

        micro_status = old[2] or 'Pending'
        if new_status.lower() == 'accepted' and micro_status.lower() == 'accepted':
            final_status = 'Accepted'
        elif 'rejected' in (new_status.lower(), micro_status.lower()):
            final_status = 'Rejected'
        else:
            final_status = 'Pending'

        cursor.execute("""
            UPDATE FinishedProductData
            SET Chemical=%s, ChemicalStatus=%s, Status=%s, Reviewed=1,
                [Reviewed By]=%s, [Reviewed Date]=%s
            WHERE [QC No.] = %s
        """, [checked, new_status, final_status, username, today, qc_no])

        log_audit(
            cursor=cursor,
            username=username,
            action_type='Edit',
            entity_type='Finished Product',
            entity_id=qc_no,
            old_values=json.dumps({'Chemical': bool(old[0]), 'ChemicalStatus': old[1], 'Status': old[3]}),
            new_values=json.dumps({'Chemical': checked, 'ChemicalStatus': new_status, 'Status': final_status}),
            details=f"Marked as reviewed (Status: {final_status})"
        )

    return Response({'message': 'Marked as reviewed'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def finished_product_mark_micro(request, qc_no):
    data = request.data
    username = data.get('username', 'Unknown')
    checked = data.get('micro', False)
    new_status = data.get('micro_status', 'Pending')
    today = date.today().isoformat()

    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT Micro, MicroStatus, ChemicalStatus, Status FROM FinishedProductData WHERE [QC No.] = %s",
            [qc_no]
        )
        old = cursor.fetchone()
        if not old:
            return Response({'error': 'Record not found'}, status=404)

        chem_status = old[2] or 'Pending'
        if new_status.lower() == 'accepted' and chem_status.lower() == 'accepted':
            final_status = 'Accepted'
        elif 'rejected' in (new_status.lower(), chem_status.lower()):
            final_status = 'Rejected'
        else:
            final_status = 'Pending'

        cursor.execute("""
            UPDATE FinishedProductData
            SET Micro=%s, MicroStatus=%s, Status=%s
            WHERE [QC No.] = %s
        """, [checked, new_status, final_status, qc_no])

        log_audit(
            cursor=cursor,
            username=username,
            action_type='Edit',
            entity_type='Finished Product',
            entity_id=qc_no,
            old_values=json.dumps({'Micro': bool(old[0]), 'MicroStatus': old[1], 'Status': old[3]}),
            new_values=json.dumps({'Micro': checked, 'MicroStatus': new_status, 'Status': final_status}),
            details=f"Mark Micro as done (Status: {final_status})"
        )

    return Response({'message': 'Micro review marked'})


# ── Raw Material Quick Marks ────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def raw_material_mark_micro(request, qc_no):
    data = request.data
    username = data.get('username', 'Unknown')
    checked = data.get('micro', False)
    new_status = data.get('micro_status', 'Pending')
    today = date.today().isoformat()

    with connection.cursor() as cursor:
        cursor.execute("SELECT Micro, MicroStatus FROM RawMaterialData WHERE [QC No.] = %s", [qc_no])
        old = cursor.fetchone()
        if not old: return Response({'error': 'Record not found'}, status=404)

        cursor.execute("""
            UPDATE RawMaterialData
            SET Micro=%s, MicroStatus=%s, [Edited By]=%s, [Edited Date]=%s
            WHERE [QC No.] = %s
        """, [checked, new_status, username, today, qc_no])

        log_audit(
            cursor=cursor, username=username, action_type='Edit', entity_type='Raw Material', entity_id=qc_no,
            old_values=json.dumps({'micro': old[0], 'micro_status': old[1]}),
            new_values=json.dumps({'micro': checked, 'micro_status': new_status})
        )

        cursor.execute("SELECT MicroStatus, ChemicalStatus FROM RawMaterialData WHERE [QC No.] = %s", [qc_no])
        statuses = cursor.fetchone()
        ms = statuses[0] or 'Pending'
        cs = statuses[1] or 'Pending'
        if ms.lower() == 'accepted' and cs.lower() == 'accepted': fs = 'Accepted'
        elif 'rejected' in (ms.lower(), cs.lower()): fs = 'Rejected'
        else: fs = 'Pending'
        cursor.execute("UPDATE RawMaterialData SET Status = %s WHERE [QC No.] = %s", [fs, qc_no])

    return Response({'message': 'Micro check updated'})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def raw_material_mark_reviewed(request, qc_no):
    data = request.data
    username = data.get('username', 'Unknown')
    checked = data.get('chemical', False)
    new_status = data.get('chemical_status', 'Pending')
    today = date.today().isoformat()

    with connection.cursor() as cursor:
        cursor.execute("SELECT Chemical, ChemicalStatus FROM RawMaterialData WHERE [QC No.] = %s", [qc_no])
        old = cursor.fetchone()
        if not old: return Response({'error': 'Record not found'}, status=404)

        cursor.execute("""
            UPDATE RawMaterialData
            SET Chemical=%s, ChemicalStatus=%s, [Edited By]=%s, [Edited Date]=%s
            WHERE [QC No.] = %s
        """, [checked, new_status, username, today, qc_no])

        log_audit(
            cursor=cursor, username=username, action_type='Edit', entity_type='Raw Material', entity_id=qc_no,
            old_values=json.dumps({'chemical': old[0], 'chemical_status': old[1]}),
            new_values=json.dumps({'chemical': checked, 'chemical_status': new_status})
        )

        cursor.execute("SELECT MicroStatus, ChemicalStatus FROM RawMaterialData WHERE [QC No.] = %s", [qc_no])
        statuses = cursor.fetchone()
        ms = statuses[0] or 'Pending'
        cs = statuses[1] or 'Pending'
        if ms.lower() == 'accepted' and cs.lower() == 'accepted': fs = 'Accepted'
        elif 'rejected' in (ms.lower(), cs.lower()): fs = 'Rejected'
        else: fs = 'Pending'
        cursor.execute("UPDATE RawMaterialData SET Status = %s WHERE [QC No.] = %s", [fs, qc_no])

    return Response({'message': 'Chemical check updated'})


# ── Packaging Material Quick Marks ────────────────────────
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def pm_mark_micro(request, qc_no):
    data = request.data
    username = data.get('username', 'Unknown')
    checked = data.get('micro', False)
    new_status = data.get('micro_status', 'Pending')

    with connection.cursor() as cursor:
        cursor.execute("SELECT Micro, MicroStatus FROM PackagingMaterialData WHERE [QC No.] = %s", [qc_no])
        old = cursor.fetchone()
        if not old: return Response({'error': 'Record not found'}, status=404)

        cursor.execute("""
            UPDATE PackagingMaterialData
            SET Micro=%s, MicroStatus=%s
            WHERE [QC No.] = %s
        """, [checked, new_status, qc_no])

        log_audit(
            cursor=cursor, username=username, action_type='Edit', entity_type='Packaging Material', entity_id=qc_no,
            old_values=json.dumps({'micro': old[0], 'micro_status': old[1]}),
            new_values=json.dumps({'micro': checked, 'micro_status': new_status})
        )

        cursor.execute("SELECT MicroStatus, ChemicalStatus FROM PackagingMaterialData WHERE [QC No.] = %s", [qc_no])
        statuses = cursor.fetchone()
        ms = statuses[0] or 'Pending'
        cs = statuses[1] or 'Pending'
        if ms.lower() == 'accepted' and cs.lower() == 'accepted': fs = 'Accepted'
        elif 'rejected' in (ms.lower(), cs.lower()): fs = 'Rejected'
        else: fs = 'Pending'
        cursor.execute("UPDATE PackagingMaterialData SET Status = %s WHERE [QC No.] = %s", [fs, qc_no])

    return Response({'message': 'Micro check updated'})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def pm_mark_reviewed(request, qc_no):
    data = request.data
    username = data.get('username', 'Unknown')
    checked = data.get('chemical', False)
    new_status = data.get('chemical_status', 'Pending')

    with connection.cursor() as cursor:
        cursor.execute("SELECT Chemical, ChemicalStatus FROM PackagingMaterialData WHERE [QC No.] = %s", [qc_no])
        old = cursor.fetchone()
        if not old: return Response({'error': 'Record not found'}, status=404)

        cursor.execute("""
            UPDATE PackagingMaterialData
            SET Chemical=%s, ChemicalStatus=%s
            WHERE [QC No.] = %s
        """, [checked, new_status, qc_no])

        log_audit(
            cursor=cursor, username=username, action_type='Edit', entity_type='Packaging Material', entity_id=qc_no,
            old_values=json.dumps({'chemical': old[0], 'chemical_status': old[1]}),
            new_values=json.dumps({'chemical': checked, 'chemical_status': new_status})
        )

        cursor.execute("SELECT MicroStatus, ChemicalStatus FROM PackagingMaterialData WHERE [QC No.] = %s", [qc_no])
        statuses = cursor.fetchone()
        ms = statuses[0] or 'Pending'
        cs = statuses[1] or 'Pending'
        if ms.lower() == 'accepted' and cs.lower() == 'accepted': fs = 'Accepted'
        elif 'rejected' in (ms.lower(), cs.lower()): fs = 'Rejected'
        else: fs = 'Pending'
        cursor.execute("UPDATE PackagingMaterialData SET Status = %s WHERE [QC No.] = %s", [fs, qc_no])

    return Response({'message': 'Chemical check updated'})

@api_view(['POST'])
<<<<<<< HEAD
<<<<<<< HEAD
@permission_classes([IsAuthenticated])
=======
@permission_classes([AllowAny])
>>>>>>> a9a5fba (Smart Log inti)
=======
@permission_classes([AllowAny])
>>>>>>> a9a5fba (Smart Log inti)
def finished_products_search(request):
    data = request.data
    current_year = datetime.now().strftime('%y')

    query = """
        SELECT Date, [Material Name], [Material Code],
               [Material Category], [Batch No.], [QC No.],
               Micro, MicroStatus, Chemical, ChemicalStatus,
               [Manufacture Date], [Expiry Date], Status,
               Notes, [Created By],
               Reviewed, [Reviewed By], [Reviewed Date]
        FROM FinishedProductData
        WHERE 1=1
    """
    params = []

    year_filter = data.get('year_filter', current_year)
    query += " AND [QC No.] LIKE %s"
    params.append(f'%{year_filter}FP')

    for field, col in [
        ('material_name', "[Material Name]"),
        ('batch_number', "[Batch No.]"),
        ('qc_number', "[QC No.]"),
        ('material_code', "[Material Code]"),
<<<<<<< HEAD
<<<<<<< HEAD
=======
        ('status', "Status"),
>>>>>>> a9a5fba (Smart Log inti)
=======
        ('status', "Status"),
>>>>>>> a9a5fba (Smart Log inti)
        ('notes', "Notes"),
    ]:
        val = data.get(field, '').strip()
        if val:
            query += f" AND {col} LIKE %s"
            params.append(f'%{val}%')

<<<<<<< HEAD
<<<<<<< HEAD
    # Exact match for status fields
    for field, col in [
        ('status', "Status"),
        ('micro_status', "MicroStatus"),
        ('chemical_status', "ChemicalStatus"),
    ]:
        val = data.get(field, '').strip()
        if val:
            query += f" AND {col} = %s"
            params.append(val)

=======
>>>>>>> a9a5fba (Smart Log inti)
=======
>>>>>>> a9a5fba (Smart Log inti)
    query += " ORDER BY [QC No.] DESC"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = dictfetchall(cursor)

    return Response({'results': rows, 'total': len(rows)})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def finished_product_generate_qc(request):
    current_year = datetime.now().strftime('%y')
    category_code = 'FP'

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT TOP 1 [QC No.]
            FROM FinishedProductData
            WHERE [Material Category] = 'Finished'
            AND [QC No.] LIKE %s
            ORDER BY [QC No.] DESC
        """, [f'%\\{current_year}{category_code}'])

        latest = cursor.fetchone()
        if latest and latest[0]:
            try:
                seq_part = latest[0].split('\\')[0]
                seq_num = int(seq_part.split('-')[0])
                new_num = f'{seq_num + 1:04d}'
            except (IndexError, ValueError):
                new_num = '0001'
        else:
            new_num = '0001'

        qc_number = f'{new_num}-00\\{current_year}{category_code}'

        cursor.execute(
            "SELECT COUNT(*) FROM FinishedProductData WHERE [QC No.] = %s",
            [qc_number]
        )
        while cursor.fetchone()[0] > 0:
            seq_num = int(new_num)
            new_num = f'{seq_num + 1:04d}'
            qc_number = f'{new_num}-00\\{current_year}{category_code}'
            cursor.execute(
                "SELECT COUNT(*) FROM FinishedProductData WHERE [QC No.] = %s",
                [qc_number]
            )

    return Response({'qc_number': qc_number})


# ──────────────────────────────────────────────
#  PRODUCTS / MATERIAL MANAGEMENT
# ──────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
<<<<<<< HEAD
<<<<<<< HEAD
def products(request): # Material Management
    with connection.cursor() as cursor:
        roles = get_user_roles(cursor, request.user.username)
        if has_any_role(roles, 'QA Manager') and not has_any_role(roles, 'Manager', 'Superuser'):
             return Response({'error': 'Unauthorized'}, status=403)
=======
def products(request):
>>>>>>> a9a5fba (Smart Log inti)
=======
def products(request):
>>>>>>> a9a5fba (Smart Log inti)
    if request.method == 'GET':
        search = request.query_params.get('search', '')
        criteria = request.query_params.get('criteria', 'ProductName')
        category = request.query_params.get('category', '')

        query = "SELECT ProductID, ProductName, ProductCategory FROM ProductNames WHERE 1=1"
        params = []

        if search:
            if criteria == 'ProductCategory':
                query += " AND ProductCategory LIKE %s"
            else:
                query += " AND ProductName LIKE %s"
            params.append(f'%{search}%')

        if category:
            query += " AND ProductCategory = %s"
            params.append(category)

        query += " ORDER BY ProductName"

        with connection.cursor() as cursor:
            cursor.execute(query, params)
            rows = dictfetchall(cursor)

        return Response({'results': rows})

    elif request.method == 'POST':
        name = request.data.get('name', '').strip()
        category = request.data.get('category', '').strip()

        if not name or not category:
            return Response({'error': 'Name and category are required'}, status=400)

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT ProductName FROM ProductNames WHERE ProductName=%s AND ProductCategory=%s",
                [name, category]
            )
            if cursor.fetchone():
                return Response({'error': 'Product already exists in this category'}, status=400)

            cursor.execute(
                "INSERT INTO ProductNames (ProductName, ProductCategory) VALUES (%s, %s)",
                [name, category]
            )

            # Audit log
            username = request.data.get('username', 'System')
            log_audit(
                cursor=cursor,
                username=username,
                action_type='Creation',
                entity_type='Product Master',
                entity_id=name,
                details=f"Category: {category}"
            )

        return Response({'message': 'Product added'}, status=201)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def product_detail(request, product_id):
    name = request.data.get('name', '').strip()
    category = request.data.get('category', '').strip()

    username = request.data.get('username', 'System')

    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT ProductID FROM ProductNames WHERE ProductName=%s AND ProductCategory=%s AND ProductID != %s",
            [name, category, product_id]
        )
        if cursor.fetchone():
            return Response({'error': 'Product name already used'}, status=400)

        cursor.execute("SELECT ProductName, ProductCategory FROM ProductNames WHERE ProductID=%s", [product_id])
        old = cursor.fetchone()

        cursor.execute(
            "UPDATE ProductNames SET ProductName=%s, ProductCategory=%s WHERE ProductID=%s",
            [name, category, product_id]
        )

        if old:
            old_name, old_category = old[0], old[1]
            if old_name != name:
                # Synchronize name across material tables
                cursor.execute("UPDATE RawMaterialData SET [Material Name] = %s WHERE [Material Name] = %s", [name, old_name])
                cursor.execute("UPDATE PackagingMaterialData SET [Material Name] = %s WHERE [Material Name] = %s", [name, old_name])
                cursor.execute("UPDATE FinishedProductData SET [Material Name] = %s WHERE [Material Name] = %s", [name, old_name])

            if old_name != name or old_category != category:
                log_audit(
                    cursor=cursor,
                    username=username,
                    action_type='Edit',
                    entity_type='Product Master',
                    entity_id=product_id,
                    old_values=json.dumps({'ProductName': old_name, 'ProductCategory': old_category}),
                    new_values=json.dumps({'ProductName': name, 'ProductCategory': category}),
                    details=f"Synchronized name change from '{old_name}' to '{name}' across material records." if old_name != name else None
                )

    return Response({'message': 'Product updated'})


# ──────────────────────────────────────────────
#  USER MANAGEMENT
# ──────────────────────────────────────────────

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def users(request):
<<<<<<< HEAD
<<<<<<< HEAD
    with connection.cursor() as cursor:
        roles = get_user_roles(cursor, request.user.username)
        if not has_any_role(roles, 'Manager'):
             return Response({'error': 'Unauthorized'}, status=403)
=======
>>>>>>> a9a5fba (Smart Log inti)
=======
>>>>>>> a9a5fba (Smart Log inti)
    if request.method == 'GET':
        with connection.cursor() as cursor:
            cursor.execute("SELECT ID, Username, Role FROM Users ORDER BY Username")
            rows = dictfetchall(cursor)
        return Response({'results': rows})

    elif request.method == 'POST':
        username = request.data.get('username', '').strip()
        password = request.data.get('password', '').strip()
        roles = request.data.get('roles', [])

        if not username or not password or not roles:
            return Response({'error': 'All fields required'}, status=400)

        hashed = bcrypt.hashpw(password.encode('utf-8'),
                               bcrypt.gensalt()).decode('utf-8')

        with connection.cursor() as cursor:
            # Enforce unique usernames at the application level
            cursor.execute("SELECT COUNT(*) FROM Users WHERE Username = %s", [username])
            if cursor.fetchone()[0] > 0:
                return Response({'error': f'Username "{username}" already exists'}, status=400)

            cursor.execute(
                "INSERT INTO Users (Username, Password, Role) VALUES (%s, %s, %s)",
                [username, hashed, roles[0]]
            )
            for r in roles[1:]:
                cursor.execute(
                    "INSERT INTO user_extra_roles (username, extra_role) VALUES (%s, %s)",
                    [username, r]
                )

            # Audit log
            actor = request.data.get('current_user', 'System')
            log_audit(
                cursor=cursor,
                username=actor,
                action_type='Creation',
                entity_type='User',
                entity_id=username,
                details=f"Roles: {', '.join(roles)}"
            )

        return Response({'message': 'User registered'}, status=201)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def user_detail(request, user_id):
    actor = request.data.get('current_user', 'System')

    with connection.cursor() as cursor:
        cursor.execute("SELECT Role, Username FROM Users WHERE ID = %s", [user_id])
        row = cursor.fetchone()
        if not row:
            return Response({'error': 'User not found'}, status=404)
        if row[0] == 'Manager':
            return Response({'error': 'Cannot delete Manager accounts'}, status=403)

        deleted_username = row[1]
        cursor.execute("DELETE FROM Users WHERE ID = %s", [user_id])

        log_audit(
            cursor=cursor,
            username=actor,
            action_type='Deletion',
            entity_type='User',
            entity_id=deleted_username,
            details=f"Deleted user with ID {user_id}"
        )

    return Response({'message': 'User deleted'})


# ──────────────────────────────────────────────
#  AUDIT TRAIL
# ──────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def audit_trail(request):
<<<<<<< HEAD
<<<<<<< HEAD
    with connection.cursor() as cursor:
        roles = get_user_roles(cursor, request.user.username)
        if not has_any_role(roles, 'Manager', 'Superuser', 'QA Manager'):
            return Response({'error': 'Unauthorized'}, status=403)
=======
>>>>>>> a9a5fba (Smart Log inti)
=======
>>>>>>> a9a5fba (Smart Log inti)
    # New unified AuditLogs query
    user_filter = request.query_params.get('user', '').strip()
    entity_type = request.query_params.get('entity_type', '').strip()
    action_type = request.query_params.get('action', '').strip()
    date_from = request.query_params.get('date_from', '').strip()
    date_to = request.query_params.get('date_to', '').strip()
    search = request.query_params.get('search', '').strip()
    offset = int(request.query_params.get('offset', 0))
    limit = int(request.query_params.get('limit', 50))

    query = """
        SELECT LogID, Timestamp, Username, ActionType, EntityType, EntityID,
               OldValues, NewValues, Details
        FROM AuditLogs
        WHERE 1=1
    """
    params = []

    if user_filter:
        query += " AND Username LIKE %s"
        params.append(f"%{user_filter}%")

    if entity_type:
        query += " AND EntityType = %s"
        params.append(entity_type)

    if action_type:
        query += " AND ActionType = %s"
        params.append(action_type)

    if date_from:
        query += " AND Timestamp >= %s"
        params.append(f"{date_from} 00:00:00")

    if date_to:
        query += " AND Timestamp <= %s"
        params.append(f"{date_to} 23:59:59")
        
    if search:
        query += " AND (EntityID LIKE %s OR Details LIKE %s)"
        params.extend([f"%{search}%", f"%{search}%"])

    query += " ORDER BY Timestamp DESC OFFSET %s ROWS FETCH NEXT %s ROWS ONLY"
    params.extend([offset, limit])

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = dictfetchall(cursor)
        
        # Get total count for pagination
        count_query = "SELECT COUNT(*) FROM AuditLogs WHERE 1=1" + query.split("WHERE 1=1")[1].split("ORDER BY")[0]
        cursor.execute(count_query, params[:-2])
        total = cursor.fetchone()[0]

    return Response({'results': rows, 'total': total})


# ──────────────────────────────────────────────
#  YEARLY REPORTS
# ──────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def yearly_report(request, category):
<<<<<<< HEAD
<<<<<<< HEAD
    with connection.cursor() as cursor:
        roles = get_user_roles(cursor, request.user.username)
        if has_any_role(roles, 'QA Manager') and not has_any_role(roles, 'Manager', 'Superuser'):
             return Response({'error': 'Unauthorized'}, status=403)
=======
>>>>>>> a9a5fba (Smart Log inti)
=======
>>>>>>> a9a5fba (Smart Log inti)
    year = request.query_params.get('year', str(datetime.now().year))
    year_suffix = year[-2:]
    offset = int(request.query_params.get('offset', 0))
    limit = int(request.query_params.get('limit', 30))
    search = request.query_params.get('search', '').strip()

    table_map = {
        'raw': ('RawMaterialData', 'RM'),
        'packaging': ('PackagingMaterialData', 'PM'),
        'finished': ('FinishedProductData', 'FP'),
    }

    if category not in table_map:
        return Response({'error': 'Invalid category'}, status=400)

    table_name, code = table_map[category]
    qc_pattern = f'%{year_suffix}{code}'
    
    where_clause = "WHERE [QC No.] LIKE %s"
    params = [qc_pattern]

    if search:
        search_parts = ["[Material Name] LIKE %s", "[QC No.] LIKE %s"]
        search_params = [f'%{search}%', f'%{search}%']
        
        if category in ('raw', 'finished'):
            search_parts.append("[Batch No.] LIKE %s")
            search_params.append(f'%{search}%')
        if category == 'packaging':
            search_parts.append("[Material Code] LIKE %s")
            search_params.append(f'%{search}%')
            
        where_clause += " AND (" + " OR ".join(search_parts) + ")"
        params.extend(search_params)

    with connection.cursor() as cursor:
        cursor.execute(f"SELECT COUNT(*) FROM {table_name} {where_clause}", params)
        total = cursor.fetchone()[0]

        cursor.execute(f"""
            SELECT * FROM {table_name}
            {where_clause}
            ORDER BY [QC No.] DESC
            OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
        """, params + [offset, limit])
        rows = dictfetchall(cursor)

    return Response({'total': total, 'results': rows})


# ──────────────────────────────────────────────
#  PRODUCT NAMES for dropdowns
# ──────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def product_names_by_category(request):
    category = request.query_params.get('category', '')
    search = request.query_params.get('search', '')

    query = "SELECT ProductName, ProductCategory FROM ProductNames WHERE 1=1"
    params = []

    if category:
        query += " AND ProductCategory = %s"
        params.append(category)
    if search:
        query += " AND ProductName LIKE %s"
        params.append(f'%{search}%')

    query += " ORDER BY ProductName"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = dictfetchall(cursor)

    return Response({'results': rows})


# ──────────────────────────────────────────────
#  SERVER TIME
# ──────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def server_time(request):
    """Return the current server datetime formatted for datetime-local inputs."""
    now = datetime.now()
    return Response({'datetime': now.strftime('%Y-%m-%dT%H:%M')})


# ──────────────────────────────────────────────
#  ADMIN HELPERS
# ──────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def drop_username_unique_constraint(request):
    """
    One-time helper: drops the FK + UNIQUE constraint on Users.Username so that
    multiple users can share the same display name.  ID is the true identifier.
    Safe to call multiple times — it does nothing if the constraints are gone.
    """
    sql = """
        -- Step 1: Drop FK on user_extra_roles that references Users.Username
        DECLARE @fk NVARCHAR(256);
        SELECT TOP 1 @fk = fk.name
        FROM sys.foreign_keys fk
        WHERE fk.parent_object_id = OBJECT_ID('user_extra_roles')
          AND fk.referenced_object_id = OBJECT_ID('Users');
        IF @fk IS NOT NULL
            EXEC('ALTER TABLE user_extra_roles DROP CONSTRAINT [' + @fk + ']');

        -- Step 2: Drop the UNIQUE constraint on Users.Username
        DECLARE @cn NVARCHAR(256);
        SELECT TOP 1 @cn = kc.name
        FROM sys.key_constraints kc
        JOIN sys.index_columns ic ON ic.object_id = kc.parent_object_id
                                  AND ic.index_id  = kc.unique_index_id
        JOIN sys.columns c        ON c.object_id   = kc.parent_object_id
                                  AND c.column_id  = ic.column_id
        WHERE kc.parent_object_id = OBJECT_ID('Users')
          AND kc.type = 'UQ'
          AND c.name  = 'Username';
        IF @cn IS NOT NULL
            EXEC('ALTER TABLE Users DROP CONSTRAINT [' + @cn + ']');
    """
    try:
        with connection.cursor() as cursor:
            cursor.execute(sql)
        return Response({'message': 'Constraints dropped (or were already absent)'})
    except Exception as e:
        return Response({'error': str(e)}, status=500)


# ──────────────────────────────────────────────
#  EXPORT DATA
# ──────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_data(request):
<<<<<<< HEAD
<<<<<<< HEAD
    with connection.cursor() as cursor:
        roles = get_user_roles(cursor, request.user.username)
        if has_any_role(roles, 'QA Manager') and not has_any_role(roles, 'Manager', 'Superuser'):
             return Response({'error': 'Unauthorized'}, status=403)
=======
>>>>>>> a9a5fba (Smart Log inti)
=======
>>>>>>> a9a5fba (Smart Log inti)
    data = request.data
    category = data.get('category', 'raw')
    year_mode = data.get('year_mode', 'specific')  # 'specific' or 'range'
    year = data.get('year')
    year_from = data.get('year_from')
    year_to = data.get('year_to')

    # Mapping category to table and base query
    if category == 'raw':
        table = "RawMaterialData"
        code_suffix = "RM"
        columns = [
            ('Date', 'Date'), ('Material Name', 'Material Name'), ('Material Code', 'Material Code'),
            ('Batch No.', 'Batch No.'), ('QC No.', 'QC No.'), ('Micro', 'Micro'),
            ('MicroStatus', 'Micro Status'), ('Chemical', 'Chemical'), ('ChemicalStatus', 'Chemical Status'),
            ('Manufacturer', 'Manufacturer'), ('Supplier', 'Supplier'), ('Manufacture Date', 'Manufacture Date'),
            ('Expiry Date', 'Expiry Date'), ('Status', 'Status'), ('Notes', 'Notes'),
            ('Created By', 'Created By'), ('Edited By', 'Edited By'), ('Edited Date', 'Edited Date')
        ]
    elif category == 'packaging':
        table = "PackagingMaterialData"
        code_suffix = "PM"
        columns = [
            ('Date', 'Date'), ('Material Name', 'Material Name'), ('Material Code', 'Material Code'),
            ('QC No.', 'QC No.'), ('Micro', 'Micro'), ('MicroStatus', 'Micro Status'),
            ('Chemical', 'Chemical'), ('ChemicalStatus', 'Chemical Status'), ('Manufacturer', 'Manufacturer'),
            ('Supplier', 'Supplier'), ('Status', 'Status'), ('Notes', 'Notes'),
            ('Created By', 'Created By'), ('Edited By', 'Edited By'), ('Edited Date', 'Edited Date')
        ]
    else:  # finished
        table = "FinishedProductData"
        code_suffix = "FP"
        columns = [
            ('Date', 'Date'), ('Material Name', 'Material Name'), ('Material Code', 'Material Code'),
            ('Batch No.', 'Batch No.'), ('QC No.', 'QC No.'), ('Micro', 'Micro'),
            ('MicroStatus', 'Micro Status'), ('Chemical', 'Chemical'), ('ChemicalStatus', 'Chemical Status'),
            ('Manufacture Date', 'Manufacture Date'), ('Expiry Date', 'Expiry Date'), ('Status', 'Status'),
            ('Notes', 'Notes'), ('Created By', 'Created By'), ('Reviewed', 'Reviewed'),
            ('Reviewed By', 'Reviewed By'), ('Reviewed Date', 'Reviewed Date')
        ]

    query = f"SELECT * FROM {table} WHERE 1=1"
    params = []

    # Year Filters
    if year_mode == 'specific' and year:
        query += " AND [QC No.] LIKE %s"
        params.append(f'%\\{str(year)[-2:]}{code_suffix}')
    elif year_mode == 'range' and year_from and year_to:
        query += " AND ("
        year_clauses = []
        for y in range(int(year_from), int(year_to) + 1):
            year_clauses.append("[QC No.] LIKE %s")
            params.append(f'%\\{str(y)[-2:]}{code_suffix}')
        query += " OR ".join(year_clauses) + ")"

    # Other Filters
    filter_map = [
        ('material_name', "[Material Name]"),
        ('qc_number', "[QC No.]"),
        ('material_code', "[Material Code]"),
        ('manufacturer', "Manufacturer"),
        ('supplier', "Supplier"),
        ('status', "Status"),
        ('micro_status', "MicroStatus"),
        ('chemical_status', "ChemicalStatus"),
    ]
    for field, col in filter_map:
        val = data.get(field, '').strip()
        if val:
            query += f" AND {col} LIKE %s"
            params.append(f'%{val}%')

    # Batch No explicitly (only for raw and finished)
    batch_val = data.get('batch_number', '').strip()
    if batch_val and category != 'packaging':
        query += " AND [Batch No.] LIKE %s"
        params.append(f'%{batch_val}%')

    if data.get('from_date'):
        query += " AND Date >= %s"
        params.append(data['from_date'])
    if data.get('to_date'):
        query += " AND Date <= %s"
        params.append(data['to_date'])

    query += " ORDER BY [QC No.] DESC"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = dictfetchall(cursor)

    # Serialize date/datetime values for JSON compatibility
    def serialize_row(row_dict):
        out = {}
        for k, v in row_dict.items():
            if isinstance(v, datetime):
                out[k] = v.strftime('%Y-%m-%d %H:%M')
            elif isinstance(v, date):
                out[k] = v.strftime('%Y-%m-%d')
            else:
                out[k] = v
        return out

    # If JSON format requested (for PDF generation on frontend)
    fmt = data.get('format', 'excel')
    if fmt == 'json':
        serialized = [serialize_row(r) for r in rows]
        return Response({'rows': serialized, 'columns': columns})

    # Generate Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"{category.replace('_', ' ').title()} Export"

    # Header
    header_font = Font(bold=True)
    for col_idx, (key, label) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font

    # Data
    for row_idx, row_dict in enumerate(rows, 2):
        for col_idx, (key, label) in enumerate(columns, 1):
            val = row_dict.get(key)
            # Format dates if needed
            if isinstance(val, (datetime, date)):
                val = val.strftime('%Y-%m-%d %H:%M') if isinstance(val, datetime) else val.strftime('%Y-%m-%d')
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Prepare response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"QC_Export_{category}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
